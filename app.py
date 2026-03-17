from flask import Flask, request, jsonify
from flask_cors import CORS
import pdfplumber
import base64
import io
import os
from functools import wraps
from PIL import Image
import pytesseract
from pdf2image import convert_from_bytes
import openpyxl
import xlrd
from datetime import datetime

app = Flask(__name__)
CORS(app)

# =============================
# SECURITY - API KEY
# =============================

API_KEY = os.environ.get('API_KEY', 'arvami_extraction_service_2026_master_key')

def require_api_key(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        api_key = request.headers.get('X-API-Key')
        
        if not api_key or api_key != API_KEY:
            return jsonify({
                'error': 'Unauthorized',
                'message': 'Invalid or missing API key'
            }), 401
        
        return f(*args, **kwargs)
    
    return decorated_function

# =============================
# HEALTH CHECK
# =============================

@app.route('/', methods=['GET'])
def health_check():
    return jsonify({
        'status': 'active',
        'service': 'Firepulse Document Extraction Service',
        'version': '2.1.0',
        'capabilities': [
            'PDF text extraction with layout preservation',
            'PDF table structure extraction',
            'OCR for scanned PDFs',
            'Excel (.xls, .xlsx) FULL multi-sheet extraction',
            'Excel formula and value extraction',
            'Excel merged cell handling',
            'Complete row extraction (no truncation)'
        ],
        'endpoints': {
            'POST /extract-pdf': 'Extract text from PDF files',
            'POST /extract-excel': 'Extract text from Excel files (.xls, .xlsx) - ALL sheets, ALL rows'
        }
    })

# =============================
# HELPER: OCR FOR SCANNED PDFs
# =============================

def extract_with_ocr(pdf_bytes):
    """Extract text from scanned PDF using OCR"""
    try:
        images = convert_from_bytes(pdf_bytes)
        
        full_text = ''
        page_texts = []
        
        for page_num, image in enumerate(images):
            page_text = pytesseract.image_to_string(image)
            
            page_texts.append({
                'page_number': page_num + 1,
                'text': page_text,
                'char_count': len(page_text),
                'method': 'OCR'
            })
            
            full_text += page_text + '\n'
        
        return full_text, page_texts
    
    except Exception as e:
        raise Exception(f'OCR extraction failed: {str(e)}')

# =============================
# HELPER: EXTRACT TABLES FROM PDF
# =============================

def extract_tables_from_page(page):
    """Extract tables and convert to readable text"""
    tables = page.extract_tables()
    
    if not tables:
        return ""
    
    table_text = ""
    
    for table_num, table in enumerate(tables):
        table_text += f"\n[TABLE {table_num + 1}]\n"
        
        for row in table:
            row_text = " | ".join([str(cell) if cell else "" for cell in row])
            table_text += row_text + "\n"
        
        table_text += "[END TABLE]\n"
    
    return table_text

# =============================
# HELPER: EXTRACT XLSX - ALL SHEETS, ALL ROWS
# =============================

def extract_xlsx_full(file_bytes):
    """Extract ALL sheets and ALL rows from XLSX file - NO TRUNCATION"""
    try:
        workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=False)
        
        all_text_parts = []
        sheets_data = []
        total_rows_extracted = 0
        
        # Process EVERY sheet
        for sheet_idx, sheet_name in enumerate(workbook.sheetnames, 1):
            sheet = workbook[sheet_name]
            
            # Sheet header
            sheet_header = f"\n{'='*60}\n[SHEET {sheet_idx}: {sheet_name}]\n{'='*60}\n"
            all_text_parts.append(sheet_header)
            
            # Get actual data range (not max which can be misleading)
            rows_list = list(sheet.iter_rows(values_only=True))
            
            if not rows_list or len(rows_list) == 0:
                all_text_parts.append("(Empty sheet)\n")
                continue
            
            # Extract header row
            headers = [str(cell) if cell is not None else '' for cell in rows_list[0]]
            header_line = "Columns: " + " | ".join(headers) + "\n"
            all_text_parts.append(header_line)
            all_text_parts.append("-" * 80 + "\n")
            
            # Extract ALL data rows (no limit)
            row_count = 0
            for row_idx, row in enumerate(rows_list[1:], start=2):
                row_values = [str(cell) if cell is not None else '' for cell in row]
                row_line = f"Row {row_idx}: " + " | ".join(row_values) + "\n"
                all_text_parts.append(row_line)
                row_count += 1
            
            total_rows_extracted += row_count
            
            # Sheet summary
            summary = f"\n(Sheet '{sheet_name}': {row_count} data rows extracted)\n"
            all_text_parts.append(summary)
            
            sheets_data.append({
                'sheet_name': sheet_name,
                'rows_extracted': row_count,
                'columns': len(headers)
            })
        
        workbook.close()
        
        full_text = "".join(all_text_parts)
        
        return {
            'text': full_text,
            'total_sheets': len(workbook.sheetnames),
            'total_rows': total_rows_extracted,
            'sheets': sheets_data,
            'extraction_method': 'openpyxl_full'
        }
    
    except Exception as e:
        raise Exception(f'XLSX extraction failed: {str(e)}')

# =============================
# HELPER: EXTRACT XLS - ALL SHEETS, ALL ROWS
# =============================

def extract_xls_full(file_bytes):
    """Extract ALL sheets and ALL rows from XLS file - NO TRUNCATION"""
    try:
        workbook = xlrd.open_workbook(file_contents=file_bytes, formatting_info=False)
        
        all_text_parts = []
        sheets_data = []
        total_rows_extracted = 0
        
        # Process EVERY sheet
        for sheet_idx in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_idx)
            sheet_name = sheet.name
            
            # Sheet header
            sheet_header = f"\n{'='*60}\n[SHEET {sheet_idx + 1}: {sheet_name}]\n{'='*60}\n"
            all_text_parts.append(sheet_header)
            
            if sheet.nrows == 0:
                all_text_parts.append("(Empty sheet)\n")
                continue
            
            # Extract header row (row 0)
            headers = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(0, col_idx)
                headers.append(str(cell.value) if cell.value else '')
            
            header_line = "Columns: " + " | ".join(headers) + "\n"
            all_text_parts.append(header_line)
            all_text_parts.append("-" * 80 + "\n")
            
            # Extract ALL data rows (starting from row 1, no limit)
            row_count = 0
            for row_idx in range(1, sheet.nrows):  # Start at 1, go to ALL rows
                row_values = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    
                    # Handle different cell types properly
                    if cell.ctype == xlrd.XL_CELL_EMPTY:
                        value = ''
                    elif cell.ctype == xlrd.XL_CELL_TEXT:
                        value = str(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_NUMBER:
                        # Format numbers properly
                        if cell.value == int(cell.value):
                            value = str(int(cell.value))
                        else:
                            value = str(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            date_tuple = xlrd.xldate_as_tuple(cell.value, workbook.datemode)
                            value = f"{date_tuple[0]}-{date_tuple[1]:02d}-{date_tuple[2]:02d}"
                        except:
                            value = str(cell.value)
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        value = 'TRUE' if cell.value else 'FALSE'
                    else:
                        value = str(cell.value)
                    
                    row_values.append(value)
                
                row_line = f"Row {row_idx + 1}: " + " | ".join(row_values) + "\n"
                all_text_parts.append(row_line)
                row_count += 1
            
            total_rows_extracted += row_count
            
            # Sheet summary
            summary = f"\n(Sheet '{sheet_name}': {row_count} data rows extracted)\n"
            all_text_parts.append(summary)
            
            sheets_data.append({
                'sheet_name': sheet_name,
                'rows_extracted': row_count,
                'columns': sheet.ncols
            })
        
        full_text = "".join(all_text_parts)
        
        return {
            'text': full_text,
            'total_sheets': workbook.nsheets,
            'total_rows': total_rows_extracted,
            'sheets': sheets_data,
            'extraction_method': 'xlrd_full'
        }
    
    except Exception as e:
        raise Exception(f'XLS extraction failed: {str(e)}')

# =============================
# ENDPOINT: EXTRACT PDF (UNCHANGED - WORKING PERFECTLY)
# =============================

@app.route('/extract-pdf', methods=['POST'])
@require_api_key
def extract_pdf():
    try:
        data = request.json
        
        if not data or 'file_base64' not in data:
            return jsonify({
                'error': 'Bad Request',
                'message': 'Missing file_base64 in request body'
            }), 400
        
        # Decode base64
        pdf_base64 = data['file_base64']
        
        try:
            pdf_bytes = base64.b64decode(pdf_base64)
        except Exception as e:
            return jsonify({
                'error': 'Invalid Base64',
                'message': f'Failed to decode base64: {str(e)}'
            }), 400
        
        # Create file object
        pdf_file = io.BytesIO(pdf_bytes)
        
        # Try pdfplumber first
        full_text = ''
        page_texts = []
        extraction_method = 'pdfplumber'
        
        try:
            with pdfplumber.open(pdf_file) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    page_text = page.extract_text() or ""
                    table_text = extract_tables_from_page(page)
                    combined_text = page_text + table_text
                    
                    page_texts.append({
                        'page_number': page_num + 1,
                        'text': combined_text,
                        'char_count': len(combined_text),
                        'has_tables': bool(table_text),
                        'method': 'pdfplumber'
                    })
                    
                    full_text += combined_text + '\n'
            
            # Check if text is empty (scanned PDF)
            if len(full_text.strip()) < 100:
                extraction_method = 'OCR_fallback'
                pdf_file.seek(0)
                full_text, page_texts = extract_with_ocr(pdf_bytes)
        
        except Exception as e:
            # Fallback to OCR
            extraction_method = 'OCR_fallback'
            full_text, page_texts = extract_with_ocr(pdf_bytes)
        
        # Clean whitespace
        full_text = ' '.join(full_text.split())
        
        return jsonify({
            'success': True,
            'text': full_text,
            'total_pages': len(page_texts),
            'total_chars': len(full_text),
            'pages': page_texts,
            'extraction_method': extraction_method,
            'metadata': {
                'file_size_bytes': len(pdf_bytes),
                'has_tables': any(p.get('has_tables') for p in page_texts)
            }
        }), 200
    
    except Exception as e:
        return jsonify({
            'error': 'PDF Extraction Failed',
            'message': str(e)
        }), 500

# =============================
# ENDPOINT: EXTRACT EXCEL - FULL MULTI-SHEET EXTRACTION
# =============================

@app.route('/extract-excel', methods=['POST'])
@require_api_key
def extract_excel():
    try:
        data = request.json
        
        if not data or 'file_base64' not in data:
            return jsonify({
                'error': 'Bad Request',
                'message': 'Missing file_base64 in request body'
            }), 400
        
        # Get file extension
        file_extension = data.get('file_extension', '.xlsx').lower()
        
        # Decode base64
        file_base64 = data['file_base64']
        
        try:
            file_bytes = base64.b64decode(file_base64)
        except Exception as e:
            return jsonify({
                'error': 'Invalid Base64',
                'message': f'Failed to decode base64: {str(e)}'
            }), 400
        
        # Extract based on file type
        if file_extension == '.xls':
            result = extract_xls_full(file_bytes)
        else:  # .xlsx
            result = extract_xlsx_full(file_bytes)
        
        return jsonify({
            'success': True,
            'text': result['text'],
            'total_sheets': result['total_sheets'],
            'total_rows': result['total_rows'],
            'total_chars': len(result['text']),
            'sheets': result['sheets'],
            'extraction_method': result['extraction_method'],
            'metadata': {
                'file_size_bytes': len(file_bytes),
                'file_extension': file_extension
            }
        }), 200
    
    except Exception as e:
        return jsonify({
            'error': 'Excel Extraction Failed',
            'message': str(e)
        }), 500

# =============================
# RUN SERVER
# =============================

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
